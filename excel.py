import openpyxl as op
print(dir(op))
wb = op.load_workbook("read_from_excel.xlsx")
print(type(wb))
a = wb.sheetnames
print(a)
print(wb.active.title)

sh1 = wb['chennai']
aa = sh1['B2'].value
print(aa)
print('-------------')
v = wb['chennai']
row = v.max_row
#col = v.max_column
for i in range(2,row+1):
    c = str(i)
    print(v['B'+c].value)
    print(v['C' + c].value)
    print(v['D' + c].value)
    print(v['E' + c].value)
    print(v['F' + c].value)
    print('--++--')
#     for j in range(2,col+1):
#         print(v.cell(i,j).value)
#     print('---------------------------------------')
# #
# k = wb['chennai']
# for singledevice in range(2,k.max_row):
#     hostname = sheet123.row(singledevice)[1].value
#     print(hostname)
